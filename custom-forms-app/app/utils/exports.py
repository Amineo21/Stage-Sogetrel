import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import json
import os
from PIL import Image as PILImage
import io
import base64
from flask import current_app, url_for
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from datetime import datetime

def export_to_excel(form_obj, responses, output_path):
    """
    Exporte les réponses d'un formulaire vers un fichier Excel
    
    Args:
        form_obj: Objet Form contenant les informations du formulaire
        responses: Liste des objets FormResponse
        output_path: Chemin du fichier Excel de sortie
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = form_obj.title[:31]  # Max 31 chars for sheet title

    # Définir les en-têtes de colonne
    headers = ['ID Réponse', 'Soumis par', 'Date de soumission', 'Adresse IP']
    field_ids = []  # Pour mapper les IDs de champ aux colonnes

    if form_obj.form_data:
        for field in form_obj.form_data:
            headers.append(field.get('label', field.get('name', field.get('id'))))
            field_ids.append(field.get('id'))

    ws.append(headers)

    # Style des en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, cell in enumerate(ws[1]):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = 20  # Largeur par défaut

    # Remplir les données
    for row_idx, response in enumerate(responses, start=2):
        username = 'Anonyme'
        if hasattr(response, 'user') and response.user:
            username = response.user.username
        elif hasattr(response, 'responder') and response.responder:
            username = response.responder.username
            
        row_data = [
            response.id,
            username,
            response.submitted_at.strftime('%Y-%m-%d %H:%M:%S'),
            response.ip_address or 'N/A'
        ]
        
        response_content = json.loads(response.response_data) if isinstance(response.response_data, str) else response.response_data
        
        for field_id in field_ids:
            field_value = response_content.get(field_id)
            
            # Trouver le type de champ pour un traitement spécifique
            field_type = None
            for field_def in form_obj.form_data:
                if field_def.get('id') == field_id:
                    field_type = field_def.get('type')
                    break

            if field_type == 'file' and field_value and isinstance(field_value, dict) and 'filename' in field_value:
                file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], field_value['filename'])
                if os.path.exists(file_path):
                    row_data.append(f"Fichier: {field_value.get('original_name', field_value['filename'])}")
                    # Optionnel: Insérer le fichier si c'est une image
                    if field_value.get('extension') in ['png', 'jpg', 'jpeg', 'gif']:
                        try:
                            img = PILImage.open(file_path)
                            # Redimensionner l'image pour qu'elle tienne dans la cellule
                            max_width = 150
                            max_height = 100
                            img.thumbnail((max_width, max_height), PILImage.Resampling.LANCZOS)
                            
                            img_byte_arr = io.BytesIO()
                            img.save(img_byte_arr, format=field_value.get('extension').upper())
                            img_byte_arr.seek(0)
                            
                            excel_img = ExcelImage(img_byte_arr)
                            # Positionner l'image dans la cellule correspondante
                            col_letter = get_column_letter(len(headers) - len(field_ids) + field_ids.index(field_id) + 1)
                            ws.add_image(excel_img, f'{col_letter}{row_idx}')
                            ws.row_dimensions[row_idx].height = max_height * 0.75  # Ajuster la hauteur de la ligne
                            ws.column_dimensions[col_letter].width = max_width / 7  # Ajuster la largeur de la colonne
                        except Exception as e:
                            current_app.logger.error(f"Could not insert image {file_path} into Excel: {e}")
                            row_data[-1] += " (Erreur d'insertion d'image)"
                else:
                    row_data.append(f"Fichier manquant: {field_value.get('original_name', field_value['filename'])}")
            elif field_type == 'signature' and field_value and isinstance(field_value, dict) and 'filename' in field_value:
                signature_path = os.path.join(current_app.config['UPLOAD_FOLDER'], field_value['filename'])
                if os.path.exists(signature_path):
                    row_data.append("Signature")
                    try:
                        img = PILImage.open(signature_path)
                        max_width = 150
                        max_height = 100
                        img.thumbnail((max_width, max_height), PILImage.Resampling.LANCZOS)
                        
                        img_byte_arr = io.BytesIO()
                        img.save(img_byte_arr, format='PNG')
                        img_byte_arr.seek(0)
                        
                        excel_img = ExcelImage(img_byte_arr)
                        col_letter = get_column_letter(len(headers) - len(field_ids) + field_ids.index(field_id) + 1)
                        ws.add_image(excel_img, f'{col_letter}{row_idx}')
                        ws.row_dimensions[row_idx].height = max_height * 0.75
                        ws.column_dimensions[col_letter].width = max_width / 7
                    except Exception as e:
                        current_app.logger.error(f"Could not insert signature {signature_path} into Excel: {e}")
                        row_data[-1] += " (Erreur d'insertion de signature)"
                else:
                    row_data.append("Signature manquante")
            elif field_type == 'checkbox':
                row_data.append('Oui' if field_value else 'Non')
            elif field_type == 'geolocation':
                if field_value:
                    row_data.append(f"Lat,Lon: {field_value}")
                else:
                    row_data.append("Non renseigné")
            else:
                row_data.append(str(field_value) if field_value is not None else '')
        
        ws.append(row_data)

    # Ajuster la largeur des colonnes automatiquement
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        if adjusted_width > 50:  # Limiter la largeur maximale pour éviter des colonnes trop larges
            adjusted_width = 50
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output_path)


def export_to_pdf(form_obj, responses):
    """
    Exporte les réponses d'un formulaire vers un fichier PDF
    
    Args:
        form_obj: Objet Form contenant les informations du formulaire
        responses: Liste des objets FormResponse
        
    Returns:
        BytesIO: Buffer contenant le PDF généré
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    
    # Container pour les éléments du PDF
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2C3E50'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#34495E'),
        spaceAfter=12
    )
    
    normal_style = styles['Normal']
    
    # Titre du formulaire
    title = Paragraph(form_obj.title, title_style)
    elements.append(title)
    
    # Description du formulaire
    if form_obj.description:
        desc = Paragraph(form_obj.description, normal_style)
        elements.append(desc)
        elements.append(Spacer(1, 12))
    
    # Informations générales
    info_text = f"<b>Nombre de réponses:</b> {len(responses)}<br/>"
    info_text += f"<b>Date de génération:</b> {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    info = Paragraph(info_text, normal_style)
    elements.append(info)
    elements.append(Spacer(1, 20))
    
    # Pour chaque réponse
    for idx, response in enumerate(responses, 1):
        # Titre de la réponse
        response_title = Paragraph(f"Réponse #{idx}", heading_style)
        elements.append(response_title)
        
        # Informations de la soumission
        username = 'Anonyme'
        if hasattr(response, 'user') and response.user:
            username = response.user.username
        elif hasattr(response, 'responder') and response.responder:
            username = response.responder.username
            
        submission_info = f"<b>Soumis par:</b> {username}<br/>"
        submission_info += f"<b>Date:</b> {response.submitted_at.strftime('%d/%m/%Y à %H:%M')}<br/>"
        submission_info += f"<b>IP:</b> {response.ip_address or 'N/A'}"
        
        info_para = Paragraph(submission_info, normal_style)
        elements.append(info_para)
        elements.append(Spacer(1, 12))
        
        # Tableau des réponses
        response_content = json.loads(response.response_data) if isinstance(response.response_data, str) else response.response_data
        
        table_data = [['Champ', 'Réponse']]
        
        for field in form_obj.form_data:
            field_id = field.get('id')
            field_label = field.get('label', field.get('name', field_id))
            field_type = field.get('type')
            field_value = response_content.get(field_id)
            
            # Traitement selon le type de champ
            if field_type == 'file' and field_value and isinstance(field_value, dict) and 'filename' in field_value:
                value_display = f"Fichier: {field_value.get('original_name', field_value['filename'])}"
            elif field_type == 'signature' and field_value and isinstance(field_value, dict) and 'filename' in field_value:
                value_display = "Signature (voir image)"
                # Ajouter l'image de signature si elle existe
                signature_path = os.path.join(current_app.config['UPLOAD_FOLDER'], field_value['filename'])
                if os.path.exists(signature_path):
                    try:
                        img = RLImage(signature_path, width=2*inch, height=1*inch)
                        table_data.append([field_label, img])
                        continue
                    except:
                        pass
            elif field_type == 'checkbox':
                value_display = 'Oui' if field_value else 'Non'
            elif field_type == 'geolocation':
                value_display = field_value if field_value else 'Non renseigné'
            else:
                value_display = str(field_value) if field_value is not None else 'Non renseigné'
            
            table_data.append([field_label, value_display])
        
        # Créer le tableau
        table = Table(table_data, colWidths=[2.5*inch, 4*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 20))
    
    # Construire le PDF
    doc.build(elements)
    
    # Retourner le buffer au début pour la lecture
    buffer.seek(0)
    return buffer
